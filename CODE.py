import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

book = openpyxl.load_workbook("C:/Users/JAHNAVI SAXENA/Desktop/Academic/Student Management Project- Python/atttendance.xlsx")
sheet = book['Sheet1']

r = sheet.max_row  # Number of rows (students)
c = sheet.max_column  # Number of columns (subjects)


resp = 1
l1 = []  # List of students to remind
l2 = ""  # Concatenated string of roll numbers with lack of attendance
l3 = []  # List of roll numbers with lack of attendance
staff_mails = ['erakshaya485@gmail.com', 'yyyyyyyy@gmail.com']  # Staff mail IDs
warning_messages = {
    1: "warning!!! you can take only one more day leave for Digital Electronic class",
    2: "warning!!! you can take only one more day leave for Maths class",
    3: "warning!!! you can take only one more day leave for Python class"
}

# Save file function
def savefile():
    book.save("C:/Users/JAHNAVI SAXENA/Desktop/Academic/Student Management Project- Python/atttendance.xlsx")
    print("saved!")

# Check attendance and send warnings
def check(no_of_days, row_num, subject_code):
    global staff_mails, l2, l3

    for student in range(0, len(row_num)):
        if no_of_days[student] == 2:
            l1.append(sheet.cell(row=row_num[student], column=2).value)
            mailstu(l1, warning_messages[subject_code])  # Send warning to student

        elif no_of_days[student] > 2:
            l2 += str(sheet.cell(row=row_num[student], column=1).value) + ","
            l3.append(sheet.cell(row=row_num[student], column=2).value)

            subject = {
                1: "Digital Electronics",
                2: "Maths",
                3: "Python"
            }[subject_code]

            msg1 = "you have lack of attendance in " + subject + " !!!"
            msg2 = "the following students have lack of attendance in your subject: " + l2[:-1]

            mailstu(l3, msg1)  # Send warning to students
            staff_id = staff_mails[subject_code - 1]  # Get respective staff's mail ID
            mailstaff(staff_id, msg2)  # Send notification to staff

# Send email to students
def mailstu(li, msg):
    from_id = 'jahnavisaxena96@gmail.com'
    pwd = 'vwzxecdbtxezvtig'

    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)

    for i in range(0, len(li)):
        to_id = li[i]
        message = MIMEMultipart()
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg, 'plain'))
        content = message.as_string()
        s.sendmail(from_id, to_id, content)

    s.quit()
    print("mail sent to students")

# Send email to staff
def mailstaff(mail_id, msg):
    from_id = 'staff0795581@gmail.com'
    pwd = 'uonbdcsuikkwwuzf'

    to_id = mail_id
    message = MIMEMultipart()
    message['Subject'] = 'Lack of attendance report'
    message.attach(MIMEText(msg, 'plain'))

    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
    content = message.as_string()
    s.sendmail(from_id, to_id, content)

    s.quit()
    print('Mail Sent to staff')

while resp == 1:
    print("1--->Digital Electronics\n2--->Maths\n3--->Python")
    y = int(input("enter subject : "))

    no_of_absentees = int(input('no.of.absentees : '))
    if no_of_absentees > 1:
        x = list(map(int, (input('roll nos :').split(' '))))
    else:
        x = [int(input('roll no : '))]

    row_num = []
    no_of_days = []

    for student in x:
        for i in range(2, r + 1):
            if y == 1:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=3).value
                    m += 1
                    sheet.cell(row=i, column=3).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)

            elif y == 2:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=4).value
                    m += 1
                    sheet.cell(row=i, column=4).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)

            elif y == 3:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=5).value
                    m += 1
                    sheet.cell(row=i, column=5).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)

    check(no_of_days, row_num, y)
    resp = int(input('another subject ? 1---->yes 0--->no: '))
